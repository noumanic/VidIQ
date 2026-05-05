"""Fill DM_Competitive_KPI.xlsx with data sourced from marketing/*.md.

Sheet 1 (Competitive Analysis): pastes Sections A–D from `10-competitive-analysis.md`.
Sheet 2 (KPI Dashboard): writes Self-Assessment (col G) + Evidence (col H) for
every row from `11-kpi-tracker.md`. Column I (Status) is left untouched — its
existing formula auto-derives from G.

Idempotent: re-running overwrites with the same values.
"""

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "DM_Competitive_KPI.xlsx"

# ── SHEET 1 — Competitive Analysis ─────────────────────────────────────────
# Each entry: (row, [VidIQ, NoteGPT, Eightify])  → cells B, C, D
COMPETITIVE_ROWS: list[tuple[int, list[str]]] = [
    # SECTION A — Company overview (rows 6–11)
    (6,  ["VidIQ",
          "NoteGPT",
          "Eightify"]),
    (7,  ["AI productivity · Video summarisation SaaS",
          "AI productivity · Video summarisation SaaS",
          "AI productivity · Chrome extension first"]),
    (8,  ["Students, educators, creators, knowledge workers, vertical-domain users (medical, legal, trading)",
          "Students, knowledge workers, casual viewers",
          "Power-YouTube viewers, students, knowledge workers"]),
    (9,  ["Multimodal — fuses transcript + vision + LLM with grounded citations",
          "Free YouTube + PDF summariser with chat",
          "YouTube summary in 8 seconds"]),
    (10, ["Free (Gemini/Groq free tier); paid tier on roadmap",
          "Freemium: 5 free summaries/day → $4.99/mo",
          "Freemium: 4 free summaries/day → $4.99/mo"]),
    (11, ["vidiq.app (to register)",
          "notegpt.io",
          "eightify.app"]),

    # SECTION B — FACEBOOK (rows 17–26)
    (17, ["(to create)",
          "facebook.com/notegpt",
          "facebook.com/eightifyapp"]),
    (18, ["0 (launch)", "~3,100", "~620"]),
    (19, ["n/a", "~12", "~4"]),
    (20, ["n/a", "~2", "<1"]),
    (21, ["7/wk planned", "~3/wk", "<1/wk"]),
    (22, ["Reels · Carousels · Single image · Stories",
          "Mostly single image + product update",
          "Mostly text + screenshot"]),
    (23, ["Confident, calm, slightly cinematic",
          "Functional, feature-led",
          "Casual, meme-adjacent"]),
    (24, ["N (setup phase)", "Y (3+ ads in Ad Library, Q4 2025)", "N"]),
    (25, ["30 s vertical + carousel + static",
          "Mostly screen-record video",
          "n/a"]),
    (26, ["\"Watch less. Learn more.\"",
          "\"Save 90% of your study time\"",
          "n/a"]),

    # SECTION B — INSTAGRAM (rows 28–35)
    (28, ["(to create)", "@notegpt.official", "@eightifyapp"]),
    (29, ["0", "~5,800", "~1,900"]),
    (30, ["n/a", "~80", "~25"]),
    (31, ["n/a", "~6", "~2"]),
    (32, ["7/wk planned", "~4/wk", "~1/wk"]),
    (33, ["Reels (35%) + Carousels + Stories",
          "Reels-heavy + Carousels",
          "Mostly static + occasional Reel"]),
    (34, ["Confident, calm",
          "Productive, hype-light",
          "Casual, dry-humour"]),
    (35, ["Y · 8–12/post (mix branded + niche)",
          "Y · ~6/post (mostly branded)",
          "Y · ~3/post"]),

    # SECTION C — Website & SEO (rows 39–44)
    (39, ["vidiq.app (to register)", "notegpt.io", "eightify.app"]),
    (40, ["1 (new domain)", "~28", "~31"]),
    (41, ["0 (new)", "~85,000", "~140,000"]),
    (42, ["aspirational: \"summarize a 2 hour youtube video\"",
          "\"youtube summarizer\"",
          "\"youtube summary chrome extension\""]),
    (43, ["aspirational: \"live stream ai summary\"",
          "\"summarize youtube video\"",
          "\"eightify\" (brand)"]),
    (44, ["aspirational: \"chat with youtube video ai\"",
          "\"youtube to text\"",
          "\"youtube ai summary\""]),

    # SECTION D — SWOT (rows 48–51): one cell per entity, multi-line
    (48, [
        "• Multimodal (vision + audio + text) unlike both\n"
        "• Live-stream pipeline (rare in this category)\n"
        "• Provider-agnostic backend with auto-failover\n"
        "• Domain-aware modes (medical, legal, trading)\n"
        "• Open-source-able (potential dev-community pull)",
        "• Established traffic, broad PDF + YT coverage\n"
        "• Free tier 5/day\n"
        "• Multilingual UI\n"
        "• Affiliate programme",
        "• Massive Chrome extension distribution (~280k DAU)\n"
        "• Free tier 4/day\n"
        "• Cleanest UX in the category\n"
        "• Strong Product Hunt history (#1 of day, Mar 2023)",
    ]),
    (49, [
        "• Zero brand awareness, zero domain authority\n"
        "• Brand-name collision with vidiq.com (YouTube SEO)\n"
        "• No mobile app\n"
        "• Small shipping team",
        "• Generic positioning, low brand recall\n"
        "• Aggressive feature creep dilutes story\n"
        "• Mobile site functional but not exceptional",
        "• Browser-locked — no mobile use case\n"
        "• Single price point, no segmentation\n"
        "• No live-stream support\n"
        "• Slow shipping cadence in 2025",
    ]),
    (50, [
        "1. Live-stream summarisation (neither competitor ships it)\n"
        "2. Vertical-domain mode (medical / legal / trading) — first-mover\n"
        "3. Pakistani + South Asian education market — low CPC, no localised competition\n"
        "4. Open-source release — fork the dev-tool audience\n"
        "5. Creator-first positioning — neither competitor markets to creators",
        # Same opportunities apply to all three players in the category — leave blank
        "(category-wide opportunities apply to all entrants)",
        "(category-wide opportunities apply to all entrants)",
    ]),
    (51, [
        "1. YouTube native AI summaries (Google rolling out auto-chapters + Gemini summaries 2025)\n"
        "2. OpenAI / Anthropic native ChatGPT-with-video features\n"
        "3. Chrome extension default (Eightify has 280k embedded users; switching cost non-zero)\n"
        "4. Free-tier API tightening (Gemini/Groq limits → unit economics flip)\n"
        "5. Brand confusion with vidiq.com (YouTube SEO tool, ~10M visits/mo)",
        "(same category-wide threats apply)",
        "(same category-wide threats apply)",
    ]),
]


# ── SHEET 2 — KPI Dashboard (rows 5–22, cols G + H) ────────────────────────
# (kpi_number, score, evidence)
KPI_ROWS: list[tuple[int, int, str]] = [
    (1,  5, "Logos in frontend/public/*.png. Typography + colour psychology written up in marketing/01-brand-guide.md §3, §5. SVG export pending."),
    (2,  5, "01-brand-guide.md §4 — full HSL/Hex table with rationale citing Mehta-Zhu (2009) and Labrecque-Milne (2012)."),
    (3,  3, "Script + storyboard + cutdowns: 02-video-ad-script.md. Footage not yet shot — KPI rises to 5 once filmed and uploaded."),
    (4,  5, "Audit table in 01-brand-guide.md §8: favicon, top-nav, hero, OG card, JSON-LD, all 4 page meta titles."),
    (5,  5, "03-content-calendar.md — 28 posts (14 days × 2 channels)."),
    (6,  5, "Same file — 4 unique formats: Reel, Carousel, Single image, Story."),
    (7,  5, "04-meta-ads-plan.md — objective ✅, budget ✅, audience (4 ad sets) ✅, creative (6 assets) ✅."),
    (8,  5, "04-meta-ads-plan.md §3 — every ad set has age, location, interests, behaviour."),
    (9,  4, "05-social-templates.md §1, §2, §3 — full templates drafted. Configure in Business Suite + screenshot per 13-self-do-checklist.md B4."),
    (10, 5, "Live deployment per 14-vercel-deployment.md (Vercel + Render free tiers). Manual click-through test per 13-self-do-checklist.md C5."),
    (11, 5, "Audit in 01-brand-guide.md §8 — logo + palette + type applied across 7 surfaces."),
    (12, 5, "Single CTA per page · TopNav navigation · Aurora hero · TanStack Query for live state · Tailwind responsive."),
    (13, 5, "All pages use Tailwind responsive classes; viewport meta correctly set. Verified in Chrome DevTools at 390 px."),
    (14, 5, "06-keyword-research.md — 18 keywords with MSV, KD, CPC."),
    (15, 5, "Same file: 5 short-tail + 13 long-tail."),
    (16, 5, "07-onpage-seo-report.md covers meta tags ✅, alt text ✅, headers ✅, KW usage ✅."),
    (17, 5, "08-google-ads-plan.md — campaign type ✅, objective ✅, audience ✅, creative ✅."),
    (18, 5, "All ad copy / creative briefs reference Pillar 1 visuals (logo + violet/fuchsia palette + display type)."),
]


def main() -> None:
    wb = load_workbook(XLSX)

    # Sheet 1 — Competitive Analysis
    ws_comp = wb["🔍 Competitive Analysis"]
    for row, (vidiq, comp1, comp2) in COMPETITIVE_ROWS:
        ws_comp.cell(row=row, column=2, value=vidiq)   # B = YOUR PRODUCT
        ws_comp.cell(row=row, column=3, value=comp1)   # C = COMPETITOR 1
        ws_comp.cell(row=row, column=4, value=comp2)   # D = COMPETITOR 2
        # Wrap text on multi-line cells (SWOT)
        if "\n" in vidiq:
            for col in (2, 3, 4):
                cell = ws_comp.cell(row=row, column=col)
                cell.alignment = cell.alignment.copy(wrapText=True, vertical="top")

    # Label NoteGPT / Eightify in the column headers (row 5, 15, 38, 47)
    for header_row in (5, 15, 38, 47):
        ws_comp.cell(row=header_row, column=3, value="NoteGPT")
        ws_comp.cell(row=header_row, column=4, value="Eightify")

    # Sheet 2 — KPI Dashboard
    ws_kpi = wb["📊 KPI Dashboard"]
    # Header row 4, KPIs occupy rows 5–22 (kpi number → row offset)
    for kpi_num, score, evidence in KPI_ROWS:
        row = 4 + kpi_num  # KPI #1 → row 5, #2 → row 6, …
        ws_kpi.cell(row=row, column=7, value=score)      # G = Self-Assessment
        ws_kpi.cell(row=row, column=8, value=evidence)   # H = Evidence
        # column I (Status) is intentionally untouched — its formula auto-derives

    wb.save(XLSX)
    print(f"✅ Wrote {len(COMPETITIVE_ROWS)} competitive rows + {len(KPI_ROWS)} KPI rows to {XLSX.name}")


if __name__ == "__main__":
    main()
