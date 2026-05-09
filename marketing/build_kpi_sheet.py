"""Generate the Competitive Analysis + KPI sheet as a .xlsx for submission.

Re-run with:
    ./venv/Scripts/python.exe marketing/build_kpi_sheet.py
Output:
    VidIQ_Submission/03_Competitive_KPI_Sheet.xlsx
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent

VIOLET = "A855F7"
FUCHSIA = "EC4899"
CYAN = "06B6D4"
EMERALD = "10B981"
AMBER = "F59E0B"
DARK = "1B142A"
LIGHT = "F4F1F8"
WHITE = "FFFFFF"


thin = Side(border_style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def header_cell(ws, row, col, text, *, fill=DARK, color=WHITE, size=11):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=size, bold=True, color=color)
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    return c


def cell(ws, row, col, text, *, bold=False, fill=None, color="1B142A", size=10,
         align="left"):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=size, bold=bold, color=color)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = border
    return c


def section_title(ws, row, text, *, span=4, fill=VIOLET):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 24


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Sheet 1: Cover ─────────────────────────────────────────────────────────


def build_cover(wb: Workbook):
    ws = wb.active
    ws.title = "Cover"
    set_widths(ws, [22, 50, 22, 22])

    # Banner
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "VidIQ — AI Video Intelligence"
    c.font = Font(name="Calibri", size=22, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42

    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = "Competitive Analysis + KPI Sheet · Digital Marketing Final Submission"
    c.font = Font(name="Calibri", size=11, color=VIOLET, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=LIGHT)
    ws.row_dimensions[2].height = 22

    # Course block
    course = [
        ("Course",        "Digital Marketing"),
        ("Section",       "CS-A"),
        ("Instructor",    "Sir Maaz Zafar Cheema"),
        ("Submission",    "Semester Course Project"),
        ("Submission date","May 2026"),
    ]
    r = 4
    for k, v in course:
        cell(ws, r, 1, k.upper(), bold=True, color="6B6080", fill=LIGHT, size=10)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        cell(ws, r, 2, v, size=11)
        r += 1

    # Group members
    r += 1
    section_title(ws, r, "GROUP MEMBERS", span=4, fill=FUCHSIA)
    r += 1
    header_cell(ws, r, 1, "Roll No.")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    header_cell(ws, r, 2, "Name")
    r += 1
    members = [
        ("22i-1653", "Insharah Aman"),
        ("21i-0416", "M. Nouman Hafeez"),
        ("21i-0484", "Shayan Khan"),
        ("21i-2507", "Muhammad Zain"),
        ("22i-1200", "Farhan Ahmed"),
    ]
    for rno, name in members:
        cell(ws, r, 1, rno, fill=LIGHT, bold=True, color=CYAN, align="center")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        cell(ws, r, 2, name)
        r += 1

    # Project links
    r += 1
    section_title(ws, r, "PROJECT LINKS", span=4, fill=CYAN)
    r += 1
    header_cell(ws, r, 1, "Surface")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    header_cell(ws, r, 2, "URL")
    r += 1
    links = [
        ("Live demo (frontend)",  "https://vidiq-two.vercel.app"),
        ("Marketing dashboard",    "https://vidiq-two.vercel.app/marketing"),
        ("Backend (Hugging Face)", "https://noumanhafeez11-vidiq-backend.hf.space"),
        ("Sitemap",                 "https://vidiq-two.vercel.app/sitemap.xml"),
        ("Robots",                  "https://vidiq-two.vercel.app/robots.txt"),
        ("Facebook Page",            "https://www.facebook.com/profile.php?id=61588939576479"),
        ("Instagram",                "https://www.instagram.com/vidiq_official/"),
    ]
    for label, url in links:
        cell(ws, r, 1, label, fill=LIGHT, bold=True, color="6B6080")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        cell(ws, r, 2, url, color="06B6D4")
        r += 1

    # Sheet index
    r += 1
    section_title(ws, r, "WORKBOOK CONTENTS", span=4, fill=EMERALD)
    r += 1
    header_cell(ws, r, 1, "Sheet")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    header_cell(ws, r, 2, "Coverage")
    r += 1
    contents = [
        ("Cover",               "Project + group + course + links + sheet index"),
        ("Competitive_Analysis","NoteGPT / Eightify side-by-side · product · social · features"),
        ("SWOT",                "Strengths · Weaknesses · Opportunities · Threats"),
        ("KPI_Tracker",         "All 18 KPIs · pillar · target · status · evidence"),
        ("KPI_Scorecard",       "Pillar-level rolled-up scores · 5 pillars"),
        ("Budget_PKR",          "₨ 250,000 line-items · paid-media split · phasing"),
        ("Flight_Targets",      "14-day campaign performance targets"),
    ]
    for s, body in contents:
        cell(ws, r, 1, s, fill=LIGHT, bold=True, color=VIOLET)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        cell(ws, r, 2, body)
        r += 1


# ── Sheet 2: Competitive Analysis ──────────────────────────────────────────


def build_competitive(wb: Workbook):
    ws = wb.create_sheet("Competitive_Analysis")
    set_widths(ws, [28, 26, 26, 26])

    section_title(ws, 1, "SECTION A — PRODUCT & PRICING", span=4, fill=VIOLET)

    headers = ["Attribute", "VidIQ", "NoteGPT", "Eightify"]
    for j, h in enumerate(headers, start=1):
        header_cell(ws, 2, j, h)

    product_rows = [
        ["Pricing",         "Free tier · $9/mo planned",    "$9/mo",            "$8.99/mo"],
        ["Free tier",       "Generous (sub-min runs)",       "10 summaries/mo",   "5 summaries/mo"],
        ["Live streams",    "Yes — rolling pipeline",        "No",                 "No"],
        ["Multimodal",      "Vision + audio + text",         "Audio + text only",  "Audio + text only"],
        ["Languages",       "11 + RTL",                       "12",                 "5"],
        ["Open-source-able","Yes",                            "No",                 "No"],
        ["Domain modes",    "Medical · legal · trading",     "No",                 "No"],
        ["Time-stamped Q&A","Yes (citations)",                "Partial",            "No"],
        ["Mobile app",      "Web-first (planned PWA)",       "iOS + Android",      "Chrome ext"],
    ]
    r = 3
    for row in product_rows:
        for j, val in enumerate(row, start=1):
            fill = LIGHT if (r % 2 == 1) else WHITE
            cell(ws, r, j, val, bold=(j == 1), fill=fill,
                 color=("6B6080" if j == 1 else "1B142A"))
        r += 1

    # Section B — Social
    r += 1
    section_title(ws, r, "SECTION B — SOCIAL PRESENCE", span=4, fill=FUCHSIA)
    r += 1
    for j, h in enumerate(headers, start=1):
        header_cell(ws, r, j, h)
    r += 1
    social_rows = [
        ["Facebook URL",     "fb.com/profile.php?id=61588939576479", "facebook.com/notegpt", "facebook.com/eightifyapp"],
        ["FB followers",     "Launching",                              "~3,100",                 "~620"],
        ["FB avg likes/post","Launching",                              "~12",                    "~4"],
        ["FB cadence",       "7 / week (planned)",                     "~3 / week",              "<1 / week"],
        ["Meta Ad activity", "Setup phase (4 sets in Review)",         "Active (Q4 2025)",       "Inactive"],
        ["FB tone",          "Confident, calm",                         "Functional, feature-led","Casual, meme-adjacent"],
        ["Instagram handle", "@vidiq_official",                         "@notegpt.official",      "@eightifyapp"],
        ["IG profile URL",   "instagram.com/vidiq_official/",           "instagram.com/notegpt.official/", "instagram.com/eightifyapp/"],
        ["IG followers",     "Launching",                                "~5,800",                 "~1,900"],
        ["IG avg likes/post","Launching",                                "~80",                    "~25"],
        ["IG cadence",       "7 / week (planned)",                      "~4 / week",               "~1 / week"],
        ["Hashtags / post",  "8–12",                                     "~6",                       "~3"],
    ]
    for row in social_rows:
        for j, val in enumerate(row, start=1):
            fill = LIGHT if (r % 2 == 1) else WHITE
            cell(ws, r, j, val, bold=(j == 1), fill=fill,
                 color=("6B6080" if j == 1 else "1B142A"))
        r += 1

    # Section C — Features matrix
    r += 1
    section_title(ws, r, "SECTION C — FEATURE COMPARISON", span=4, fill=CYAN)
    r += 1
    for j, h in enumerate(headers, start=1):
        header_cell(ws, r, j, h)
    r += 1
    features = [
        ["YouTube videos",      "✅", "✅", "✅"],
        ["Live streams",        "✅", "❌", "❌"],
        ["Vision / keyframes",  "✅", "❌", "❌"],
        ["Time-stamped chat",   "✅", "Partial", "❌"],
        ["Translation",         "✅ 11 langs + RTL", "✅ 12 langs", "✅ 5 langs"],
        ["Comparison view",     "✅", "❌", "❌"],
        ["Analytics dashboard", "✅", "❌", "❌"],
        ["Chrome extension",    "Planned", "❌", "✅"],
        ["Mobile app",          "Planned PWA", "✅", "❌"],
        ["API access",          "Planned",     "✅", "❌"],
    ]
    for row in features:
        for j, val in enumerate(row, start=1):
            fill = LIGHT if (r % 2 == 1) else WHITE
            cell(ws, r, j, val, bold=(j == 1), fill=fill,
                 color=("6B6080" if j == 1 else "1B142A"),
                 align=("left" if j == 1 else "center"))
        r += 1

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "B3"


# ── Sheet 3: SWOT ──────────────────────────────────────────────────────────


def build_swot(wb: Workbook):
    ws = wb.create_sheet("SWOT")
    set_widths(ws, [55, 55])

    section_title(ws, 1, "SWOT ANALYSIS — VidIQ", span=2, fill=DARK)

    quads = [
        ("STRENGTHS", EMERALD, [
            "Multimodal AI: vision + audio + text fused (neither competitor)",
            "Live-stream pipeline — rare in this category",
            "Provider-agnostic backend with auto-failover (Gemini ↔ Groq ↔ OpenAI)",
            "Domain-aware modes (medical · legal · trading · education)",
            "Open-source-able product — potential dev-community pull",
            "Time-stamped citations behind every chat answer",
            "Pakistani CPM/CPC arbitrage: 5–7× cheaper reach per rupee",
        ]),
        ("WEAKNESSES", "F43F5E", [
            "Zero brand awareness · DA = 1",
            "Brand-name collision with vidiq.com (unrelated YouTube SEO tool)",
            "No native mobile app yet (web-only)",
            "Small shipping team (5 students)",
            "YouTube IP rate-limiting on the HF Space demo",
            "Dependence on free-tier LLM quotas in showcase mode",
        ]),
        ("OPPORTUNITIES", CYAN, [
            "Long-tail SEO is rankable in 2–3 months (KD < 15)",
            "First-mover in vertical niches (trading streams, med lectures, legal)",
            "Pakistani student / creator market underserved by AI tools",
            "Creator economy demand for time-saving tooling is growing",
            "Free-tier-first SaaS positioning vs premium-only competitors",
            "Translation feature targets Urdu / Arabic / RTL audiences",
        ]),
        ("THREATS", AMBER, [
            "Big-tech could ship native summaries (YouTube, Google Bard)",
            "LLM API pricing changes could compress margins",
            "YouTube ToS changes around transcript scraping",
            "Brand-name confusion with the existing vidiq.com property",
            "Meta / Google ad policy changes in Pakistan",
            "Currency volatility (PKR/USD) affects ad budget purchasing power",
        ]),
    ]

    # 2x2 grid layout
    positions = [(2, 1), (2, 2), (12, 1), (12, 2)]
    for (r0, c0), (label, color, items) in zip(positions, quads):
        # quadrant header
        h = ws.cell(row=r0, column=c0, value=label)
        h.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
        h.fill = PatternFill("solid", fgColor=color)
        h.alignment = Alignment(horizontal="center", vertical="center")
        h.border = border
        ws.row_dimensions[r0].height = 22
        # items
        for i, item in enumerate(items, start=1):
            c = ws.cell(row=r0 + i, column=c0, value=f"• {item}")
            c.font = Font(name="Calibri", size=10, color="1B142A")
            c.fill = PatternFill("solid", fgColor=(LIGHT if i % 2 == 1 else WHITE))
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            c.border = border
            ws.row_dimensions[r0 + i].height = 24


# ── Sheet 4: KPI Tracker ───────────────────────────────────────────────────


def build_kpi_tracker(wb: Workbook):
    ws = wb.create_sheet("KPI_Tracker")
    set_widths(ws, [4, 14, 36, 25, 18, 32])

    section_title(ws, 1, "KPI TRACKER — 18 GRADED KPIs", span=6, fill=DARK)

    headers = ["#", "Pillar", "KPI", "Target / Threshold", "Status", "Evidence / Source"]
    for j, h in enumerate(headers, start=1):
        header_cell(ws, 2, j, h)

    kpis = [
        ("1",  "Pillar 1 — Brand",   "Logo applied across all surfaces",       "8 surfaces · consistent",     "Strong",
         "Slide 33 · favicon · OG card · FB cover · IG profile"),
        ("2",  "Pillar 1 — Brand",   "Brand voice documented",                  "5 values + tone slider",     "Strong",
         "01-brand-guide.md · slide 5"),
        ("3",  "Pillar 1 — Brand",   "Colour & typography system",              "6 tokens + 3 typefaces",     "Strong",
         "globals.css · slide 21 · build_deck.py"),
        ("4",  "Pillar 1 — Brand",   "Brand consistency across deliverables",   "5 / 5 surfaces",              "Strong",
         "Slide 33 · audit · 7/7 surfaces green"),
        ("5",  "Pillar 2 — Social",  "14-day social calendar",                  "14 / 14 days posted",        "Strong",
         "Business Suite Planner screenshot · 03-content-calendar.md"),
        ("6",  "Pillar 2 — Social",  "Format mix (Reel · Carousel · Image · Story)","4 of 4 formats",        "Strong",
         "05-social-templates.md · planner-calendar.png"),
        ("7",  "Pillar 2 — Social",  "FB + IG identities live",                 "Both verified",               "Strong",
         "fb.com/…?id=61588939576479 · @vidiq_official"),
        ("8",  "Pillar 2 — Social",  "Meta Ads — 4 ad sets to Review",          "4 / 4 built",                 "Strong",
         "meta-ad-set-{1-4}-*.png · slide 8"),
        ("9",  "Pillar 3 — Convo",   "Welcome message + auto-FAQ",              "Configured + tested",         "Strong",
         "auto-greeting.png · auto-faq.png"),
        ("10", "Pillar 3 — Convo",   "Saved replies + auto-away",                "8 replies + after-hours",     "Strong",
         "saved-replies.png · auto-away.png"),
        ("11", "Pillar 3 — Convo",   "Tone consistent across human + bot",      "Voice rubric pass",           "Strong",
         "01-brand-guide.md voice slider"),
        ("12", "Pillar 4 — SEO",     "Keyword research · 18 KWs · 3 tiers",     "5 head + 13 long-tail",       "Strong",
         "06-keyword-research.md · slide 26"),
        ("13", "Pillar 4 — SEO",     "On-page SEO · 8 techniques applied",      "8 / 8 implemented",           "Strong",
         "07-onpage-seo-report.md · slide 11"),
        ("14", "Pillar 4 — SEO",     "GSC verified + sitemap submitted",        "Property + sitemap live",     "Strong",
         "gsc-property-verified.png · gsc-sitemap-success.png"),
        ("15", "Pillar 4 — Ads",     "Google Ads · 3 campaigns to Review",      "Search + YouTube + PMax",     "Strong",
         "google-{search,youtube,pmax}-review.* · slide 12"),
        ("16", "Pillar 4 — Ads",     "Bidding strategy + guardrails",            "CPA cap + neg-KW sweep",      "Strong",
         "08-google-ads-plan.md · slide 27"),
        ("17", "Pillar 5 — P5",      "Total budget + distribution (PKR)",       "₨ 250,000 plan",              "Strong",
         "09-budget.md · slides 13 + 28"),
        ("18", "Pillar 5 — P5",      "Competitive analysis · 2 competitors",     "NoteGPT + Eightify",          "Strong",
         "10-competitive-analysis.md · slides 14 + 15"),
    ]
    r = 3
    for row in kpis:
        for j, val in enumerate(row, start=1):
            fill = LIGHT if (r % 2 == 1) else WHITE
            text_color = "1B142A"
            if j == 5:  # Status
                text_color = "10B981" if val == "Strong" else "F59E0B"
            elif j == 1:
                text_color = "A855F7"
            cell(ws, r, j, val,
                 bold=(j in (1, 5)),
                 fill=fill, color=text_color,
                 align=("center" if j in (1, 5) else "left"))
        r += 1

    ws.freeze_panes = "A3"


# ── Sheet 5: KPI Scorecard ────────────────────────────────────────────────


def build_scorecard(wb: Workbook):
    ws = wb.create_sheet("KPI_Scorecard")
    set_widths(ws, [22, 12, 14, 14, 32])

    section_title(ws, 1, "PILLAR-LEVEL SCORECARD", span=5, fill=DARK)

    headers = ["Pillar", "KPIs", "Met", "Score", "Notes"]
    for j, h in enumerate(headers, start=1):
        header_cell(ws, 2, j, h)

    pillars = [
        ("Pillar 1 — Branding",       4, 4, "3 / 3", "All 4 brand KPIs met (logo, voice, colour, consistency)"),
        ("Pillar 2 — Social",          4, 4, "3 / 3", "Calendar, formats, profiles, Meta ads — all green"),
        ("Pillar 3 — Conversation",    3, 3, "3 / 3", "Welcome, FAQ, saved replies, auto-away — all configured"),
        ("Pillar 4 — SEO + Ads",       5, 5, "3 / 3", "18 KWs, 8 on-page techniques, GSC verified, 3 Google campaigns"),
        ("Pillar 5 — Competitive/KPI", 2, 2, "3 / 3", "PKR budget plan + NoteGPT/Eightify competitive analysis"),
    ]
    r = 3
    for row in pillars:
        for j, val in enumerate(row, start=1):
            fill = LIGHT if (r % 2 == 1) else WHITE
            color = "10B981" if j == 4 else "1B142A"
            cell(ws, r, j, val,
                 bold=(j in (1, 4)),
                 fill=fill, color=color,
                 align=("center" if j in (2, 3, 4) else "left"))
        r += 1

    # Total row
    cell(ws, r, 1, "TOTAL", bold=True, fill=DARK, color=WHITE)
    cell(ws, r, 2, "18", bold=True, fill=DARK, color=WHITE, align="center")
    cell(ws, r, 3, "18", bold=True, fill=DARK, color=WHITE, align="center")
    cell(ws, r, 4, "15 / 15", bold=True, fill=EMERALD, color=WHITE, align="center")
    cell(ws, r, 5, "All five pillar deliverables complete · ready for viva", bold=True, fill=DARK, color=WHITE)
    ws.row_dimensions[r].height = 24

    # Note
    r += 2
    section_title(ws, r, "ADDITIONAL · PRESENTATION & Q&A (10 marks)", span=5, fill=FUCHSIA)
    r += 1
    notes = [
        ("Coverage",      "33-slide deck (15 main + appendix) + final report in .docx"),
        ("Defense",       "Each KPI links to verifiable artefact (URL or screenshot)"),
        ("Coordination",  "5 members · roles documented on cover · all sections rehearsed"),
        ("Live demo",     "vidiq-two.vercel.app — public, indexed, working"),
    ]
    for k, v in notes:
        cell(ws, r, 1, k, bold=True, fill=LIGHT, color="6B6080")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        cell(ws, r, 2, v)
        r += 1


# ── Sheet 6: Budget ────────────────────────────────────────────────────────


def build_budget(wb: Workbook):
    ws = wb.create_sheet("Budget_PKR")
    set_widths(ws, [28, 16, 14, 12, 36])

    section_title(ws, 1, "BUDGET — PKR-PRIMARY · ₨ 250,000 / 14-DAY FLIGHT", span=5, fill=DARK)

    headers = ["Line item", "PKR", "USD (~280 PKR)", "Share", "Notes"]
    for j, h in enumerate(headers, start=1):
        header_cell(ws, 2, j, h)

    items = [
        ("Meta Ads",            "₨ 95,000",  "$340",  "38 %", "4 ad sets · ₨ 23,750 each (Students · Creators · Knowledge · Retarget)"),
        ("Google Ads",          "₨ 70,000",  "$250",  "28 %", "3 campaigns: Search ₨ 23,100 · YouTube ₨ 18,900 · PMax ₨ 28,000"),
        ("Creative production", "₨ 40,000",  "$143",  "16 %", "30s ad master · 4 cutdowns · brand assets · stills"),
        ("Influencer seeding",  "₨ 20,000",  "$71",    "8 %", "3–5 micro-creators · gifted access · UGC rights"),
        ("Tools & SaaS",         "₨ 17,500",  "$63",    "7 %", "Ubersuggest pro · Canva pro · scheduling tools (free tiers cover ~80%)"),
        ("Contingency (10%)",    "₨ 7,500",   "$27",    "3 %", "Reserved for fatigue rotation + creative refreshes"),
    ]
    r = 3
    for row in items:
        for j, val in enumerate(row, start=1):
            fill = LIGHT if (r % 2 == 1) else WHITE
            cell(ws, r, j, val, bold=(j == 1), fill=fill,
                 color=("6B6080" if j == 1 else "1B142A"),
                 align=("right" if j in (2, 3, 4) else "left"))
        r += 1

    # Total
    cell(ws, r, 1, "TOTAL", bold=True, fill=DARK, color=WHITE)
    cell(ws, r, 2, "₨ 250,000", bold=True, fill=DARK, color=WHITE, align="right")
    cell(ws, r, 3, "$893",      bold=True, fill=DARK, color=WHITE, align="right")
    cell(ws, r, 4, "100 %",     bold=True, fill=DARK, color=WHITE, align="right")
    cell(ws, r, 5, "14-day flight · ₨ 0 actual (showcase / Review-only)", bold=True, fill=DARK, color=WHITE)
    ws.row_dimensions[r].height = 22

    # Rationale
    r += 2
    section_title(ws, r, "WHY ₨ 250,000?", span=5, fill=AMBER)
    r += 1
    rationale = [
        ("Market norm",   "Pakistani SaaS launch budgets cluster at ₨ 150K–400K/month — ₨ 250K is mid-range, aggressive for a 14-day flight."),
        ("Meta CPM (PK)",  "₨ 80–250 per 1,000 impressions (vs $5–15 in US/UK) — 5–7× more reach per rupee."),
        ("Google CPC (PK)","Long-tail KWs: ₨ 20–80 per click (vs $1–3 USD) — same order-of-magnitude gap."),
        ("Volume bought",  "₨ 250K ≈ 80,000–120,000 impressions across Meta + Google at our PK targeting."),
        ("Affordability",  "Sized for a final-year student team without external sponsorship."),
        ("Pixel learning", "Spend is enough for the Meta pixel to leave the Learning phase and converge on a stable CPA before D8."),
    ]
    for k, v in rationale:
        cell(ws, r, 1, k, bold=True, fill=LIGHT, color="A855F7")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        cell(ws, r, 2, v)
        r += 1

    # Phasing
    r += 1
    section_title(ws, r, "SPEND PHASING — 14 DAYS", span=5, fill=CYAN)
    r += 1
    phase_headers = ["Window", "Phase", "Action", "", ""]
    for j, h in enumerate(phase_headers[:3], start=1):
        header_cell(ws, r, j, h)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    r += 1
    phases = [
        ("Days 1–7",  "Learn",              "Highest-Volume bid · LP-views objective · cheapest data while pixel learns"),
        ("Day 5",     "Creative-swap fund", "10 % of Meta budget reserved for fatigue rotation"),
        ("Days 8–14", "Optimise",           "Switch to Conversions · CPR cap ₨ 420 · PMax unlocks if 30 conv banked"),
        ("Day 14",    "Wrap",               "Negative-KW sweep · screenshot reports · publish learnings"),
    ]
    for window, phase, action in phases:
        fill = LIGHT if (r % 2 == 1) else WHITE
        cell(ws, r, 1, window, bold=True, fill=fill, color="6B6080")
        cell(ws, r, 2, phase, bold=True, fill=fill, color="A855F7")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        cell(ws, r, 3, action, fill=fill)
        r += 1


# ── Sheet 7: Flight Targets ───────────────────────────────────────────────


def build_targets(wb: Workbook):
    ws = wb.create_sheet("Flight_Targets")
    set_widths(ws, [30, 22, 22, 38])

    section_title(ws, 1, "14-DAY FLIGHT — PERFORMANCE TARGETS", span=4, fill=DARK)

    headers = ["KPI", "Target (PKR / %)", "Target (USD)", "Source / Benchmark"]
    for j, h in enumerate(headers, start=1):
        header_cell(ws, 2, j, h)

    rows = [
        ("Cost per result (Meta)",       "≤ ₨ 420",    "≤ $1.50",   "Ads Manager · LP-views → Conversions ladder"),
        ("Cost per click (Google Search)","≤ ₨ 70",     "≤ $0.25",   "Industry benchmark for high-intent KWs"),
        ("CTR (Meta)",                    "≥ 1.2 %",    "—",          "AI-tools vertical benchmark"),
        ("CTR (Google Search)",           "≥ 4 %",      "—",          "High-intent KW benchmark"),
        ("Demo starts / day",              "30",         "—",          "GA4 · custom event start_analysis"),
        ("Sign-ups / day",                 "8",          "—",          "GA4 · custom event sign_up"),
        ("Frequency (Meta)",               "≤ 2.5",      "—",          "Avoids audience fatigue in 14d window"),
        ("ROAS (Meta + Google blended)",   "≥ 1.4× by D14","—",       "Conversions × $9 LTV proxy"),
        ("Pixel events / day",             "≥ 50",       "—",          "Conversions API + browser pixel"),
        ("Sessions / day (organic + paid)", "200+ by D7", "—",          "GA4 · landing-page sessions"),
    ]
    r = 3
    for row in rows:
        for j, val in enumerate(row, start=1):
            fill = LIGHT if (r % 2 == 1) else WHITE
            cell(ws, r, j, val,
                 bold=(j == 1), fill=fill,
                 color=("1B142A" if j == 1 else "10B981" if j in (2, 3) else "6B6080"),
                 align=("left" if j in (1, 4) else "center"))
        r += 1

    # Tracking architecture
    r += 1
    section_title(ws, r, "CONVERSION-TRACKING ARCHITECTURE", span=4, fill=FUCHSIA)
    r += 1
    rows2 = [
        ("Meta Pixel + CAPI", "Browser pixel + server-side Conversions API · same event names"),
        ("Google gtag",        "Site-wide gtag.js · auto-tagged for cross-device measurement"),
        ("GA4",                "Property linked to GSC · enhanced measurement on"),
        ("Event taxonomy",     "view_demo · start_analysis · finish_analysis · open_chat · sign_up"),
        ("Attribution",        "Data-driven (Google) + 7-day click / 1-day view (Meta)"),
    ]
    for k, v in rows2:
        cell(ws, r, 1, k, bold=True, fill=LIGHT, color="A855F7")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        cell(ws, r, 2, v)
        r += 1


# ── Build ─────────────────────────────────────────────────────────────────


def main():
    wb = Workbook()
    build_cover(wb)
    build_competitive(wb)
    build_swot(wb)
    build_kpi_tracker(wb)
    build_scorecard(wb)
    build_budget(wb)
    build_targets(wb)

    out_dir = ROOT / "VidIQ_Submission"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / "03_Competitive_KPI_Sheet.xlsx"
    try:
        wb.save(out_file)
    except PermissionError:
        i = 2
        while (out_dir / f"03_Competitive_KPI_Sheet_v{i}.xlsx").exists():
            i += 1
        out_file = out_dir / f"03_Competitive_KPI_Sheet_v{i}.xlsx"
        wb.save(out_file)
        print(f"[warn] primary file locked — wrote {out_file.name}")
    print(f"[ok] wrote {out_file}")


if __name__ == "__main__":
    main()
